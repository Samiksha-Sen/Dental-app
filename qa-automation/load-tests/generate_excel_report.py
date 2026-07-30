"""Generates Load_Testing_Report.xlsx from load-test-summary.json so that
downloading load-test-reports artifact in GitHub Actions gives a clean Excel spreadsheet.
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def build_load_excel_report():
    reports_dir = Path(__file__).resolve().parent / "reports"
    summary_json = reports_dir / "load-test-summary.json"
    output_path = reports_dir / "Load_Testing_Report.xlsx"

    if not summary_json.exists():
        print(f"Warning: {summary_json} not found. Creating empty load report.", file=sys.stderr)
        results = []
    else:
        results = json.loads(summary_json.read_text(encoding="utf-8"))

    total = len(results)
    passed = sum(1 for r in results if r.get("status") == "Passed")
    failed = total - passed

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = [
        "Execution Date",
        "Environment",
        "Total Scenarios",
        "Passed",
        "Failed",
        "Pass %",
    ]
    ws_summary.append(summary_headers)
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    pass_pct = f"{(passed / total * 100):.1f}%" if total > 0 else "N/A"
    ws_summary.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Flask AI API (Local CI)",
        total,
        passed,
        failed,
        pass_pct,
    ])
    for cell in ws_summary[2]:
        cell.alignment = center_align

    # Sheet 2: Scenario Results
    ws_scenarios = wb.create_sheet(title="Scenario Results")
    scenario_headers = [
        "Scenario ID",
        "Status",
        "p95 Latency (ms)",
        "Error Rate (%)",
        "Request Count",
        "Failure Count",
    ]
    ws_scenarios.append(scenario_headers)
    for col_idx in range(1, len(scenario_headers) + 1):
        cell = ws_scenarios.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for r in results:
        ws_scenarios.append([
            r.get("id", ""),
            r.get("status", ""),
            r.get("p95Ms", 0),
            r.get("errorRatePct", 0),
            r.get("requestCount", 0),
            r.get("failureCount", 0),
        ])

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    reports_dir.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Load testing Excel report written to {output_path} ({total} scenarios, {passed} passed)")


if __name__ == "__main__":
    build_load_excel_report()
