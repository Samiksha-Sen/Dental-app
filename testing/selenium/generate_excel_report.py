"""Generates Selenium_E2E_Test_Report.xlsx from pytest results.xml so that
downloading the selenium-results artifact in GitHub Actions gives a clean Excel spreadsheet.
"""

import os
import sys
import xml.etree.ElementTree as ET
from datetime import datetime

import openpyxl
from pathlib import Path
from openpyxl.styles import Alignment, Font, PatternFill


def build_excel_report(xml_path=None, output_path=None):
    script_dir = Path(__file__).resolve().parent
    xml_path = Path(xml_path) if xml_path else script_dir / "results.xml"
    output_path = Path(output_path) if output_path else script_dir / "Selenium_E2E_Test_Report.xlsx"

    if not os.path.exists(xml_path):
        print(f"Warning: {xml_path} not found. Creating an empty summary report.", file=sys.stderr)
        total, passed, failed, skipped, duration = 0, 0, 0, 0, 0.0
        test_cases = []
    else:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        testsuite = root if root.tag == "testsuite" else root.find("testsuite")
        if testsuite is None and len(root) > 0:
            testsuite = root[0]

        total = int(testsuite.attrib.get("tests", 0)) if testsuite is not None else 0
        failed = int(testsuite.attrib.get("failures", 0)) if testsuite is not None else 0
        skipped = int(testsuite.attrib.get("skipped", 0)) if testsuite is not None else 0
        errors = int(testsuite.attrib.get("errors", 0)) if testsuite is not None else 0
        duration = float(testsuite.attrib.get("time", 0.0)) if testsuite is not None else 0.0

        passed = max(0, total - failed - skipped - errors)
        failed += errors

        test_cases = []
        cases_elem = testsuite if testsuite is not None else []
        for idx, case in enumerate(cases_elem.findall("testcase"), start=1):
            classname = case.attrib.get("classname", "Selenium")
            name = case.attrib.get("name", "Test Case")
            time_taken = float(case.attrib.get("time", 0.0))

            failure = case.find("failure")
            error = case.find("error")
            skip = case.find("skipped")

            if failure is not None:
                status = "Failed"
                message = failure.attrib.get("message") or failure.text or "Test Failed"
            elif error is not None:
                status = "Failed"
                message = error.attrib.get("message") or error.text or "Test Error"
            elif skip is not None:
                status = "Skipped"
                message = skip.attrib.get("message") or skip.text or "Skipped"
            else:
                status = "Passed"
                message = ""

            test_cases.append({
                "id": f"SEL-{str(idx).zfill(3)}",
                "module": classname.split(".")[-1],
                "scenario": name,
                "status": status,
                "duration": round(time_taken, 2),
                "message": str(message).strip() if message else "",
            })

    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = [
        "Execution Date",
        "Environment",
        "Total Tests",
        "Passed",
        "Failed",
        "Skipped",
        "Pass %",
        "Duration (s)",
    ]
    ws_summary.append(summary_headers)
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    pass_pct = f"{(passed / total * 100):.1f}%" if total > 0 else "N/A"
    ws_summary.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Selenium Chrome (Web)",
        total,
        passed,
        failed,
        skipped,
        pass_pct,
        f"{duration:.2f}",
    ])
    for cell in ws_summary[2]:
        cell.alignment = center_align

    # Sheet 2: Test Cases
    ws_cases = wb.create_sheet(title="Test Cases")
    case_headers = [
        "Test ID",
        "Module",
        "Scenario / Test Name",
        "Status",
        "Duration (s)",
        "Failure Message",
    ]
    ws_cases.append(case_headers)
    for col_idx in range(1, len(case_headers) + 1):
        cell = ws_cases.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for tc in test_cases:
        ws_cases.append([
            tc["id"],
            tc["module"],
            tc["scenario"],
            tc["status"],
            tc["duration"],
            tc["message"],
        ])

    # Sheet 3: Failed Tests (if any)
    failed_cases = [tc for tc in test_cases if tc["status"] == "Failed"]
    if failed_cases:
        ws_failed = wb.create_sheet(title="Failed Tests")
        fail_headers = ["Test ID", "Scenario / Test Name", "Failure Details"]
        ws_failed.append(fail_headers)
        for col_idx in range(1, len(fail_headers) + 1):
            cell = ws_failed.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        for fc in failed_cases:
            ws_failed.append([fc["id"], fc["scenario"], fc["message"]])

    # Auto-fit columns
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 75)

    wb.save(output_path)
    print(f"Excel report successfully written to {output_path} ({total} total, {passed} passed, {failed} failed)")


if __name__ == "__main__":
    xml_file = sys.argv[1] if len(sys.argv) > 1 else None
    out_file = sys.argv[2] if len(sys.argv) > 2 else None
    build_excel_report(xml_file, out_file)
